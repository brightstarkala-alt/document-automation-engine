import re
import shutil
import subprocess
import tempfile
from io import BytesIO
from pathlib import Path

import openpyxl
from openpyxl.cell.cell import MergedCell

PLACEHOLDER_RE = re.compile(r"\{\{(\w+)\}\}")


def _to_xlsx(excel_bytes: bytes, suffix: str) -> bytes:
    """If the file is old .xls format, convert to .xlsx via LibreOffice."""
    if suffix.lower() == ".xlsx":
        return excel_bytes
    with tempfile.TemporaryDirectory() as tmpdir:
        tmp = Path(tmpdir)
        src = tmp / f"form{suffix}"
        src.write_bytes(excel_bytes)
        soffice = (
            shutil.which("soffice")
            or "/Applications/LibreOffice.app/Contents/MacOS/soffice"
        )
        subprocess.run(
            [soffice, "--headless", "--convert-to", "xlsx", "--outdir", str(tmp), str(src)],
            check=True,
            capture_output=True,
        )
        return (tmp / "form.xlsx").read_bytes()


def _resolve_cell(ws, row: int, col: int):
    """Return the writable master cell, resolving MergedCell references."""
    cell = ws.cell(row=row, column=col)
    if not isinstance(cell, MergedCell):
        return cell
    for rng in ws.merged_cells.ranges:
        if rng.min_row <= row <= rng.max_row and rng.min_col <= col <= rng.max_col:
            return ws.cell(row=rng.min_row, column=rng.min_col)
    return cell


def _make_field_name(text: str) -> str:
    text = re.sub(r"[^a-z0-9\s]", "", text.lower())
    text = re.sub(r"\s+", "_", text.strip())
    return text[:50] or "field"


def _is_label(text: str) -> bool:
    """Return True only if text looks like a form label, not pre-filled content."""
    text = text.strip()
    if not text:
        return False
    # ALL-CAPS strings are almost always pre-filled values (company names, addresses, codes)
    if text == text.upper() and any(c.isalpha() for c in text):
        return False
    # More than 2 consecutive digits → likely a reference number, date, or pre-filled code
    if re.search(r"\d{3,}", text):
        return False
    # More than 6 words → too long to be a label
    if len(text.split()) > 6:
        return False
    return True


def _cell_ref(row: int, col: int) -> str:
    """Convert row/col to Excel-style cell reference (e.g. B5)."""
    col_str = ""
    c = col
    while c > 0:
        c, rem = divmod(c - 1, 26)
        col_str = chr(ord("A") + rem) + col_str
    return f"{col_str}{row}"


def _label_from_position(ws, row: int, col: int) -> str:
    if col <= 1:
        return f"field {row}_{col}"
    left = ws.cell(row=row, column=col - 1)
    if left.value and isinstance(left.value, str):
        return left.value.strip().rstrip(":")
    return f"field {row}_{col}"


def build_excel_preview_metadata(
    excel_bytes: bytes,
    field_positions: list[dict],
    suffix: str = ".xlsx",
) -> dict:
    """Build grid-based preview metadata preserving sheet layout."""
    wb = openpyxl.load_workbook(BytesIO(_to_xlsx(excel_bytes, suffix)))
    ws = wb.active

    regions = []
    for pos in field_positions:
        row, col = pos["row"], pos["col"]
        regions.append({
            "field_id": pos["name"],
            "variable_name": pos["name"],
            "label_text": pos.get("label", pos["name"]),
            "cell_ref": _cell_ref(row, col),
            "row": row,
            "col": col,
            "source": "excel",
        })

    return {
        "version": 1,
        "page_count": 1,
        "file_type": "excel",
        "sheet_name": ws.title,
        "max_row": ws.max_row,
        "max_col": ws.max_column,
        "pages": [{
            "page_index": 0,
            "width_px": 0,
            "height_px": 0,
            "thumbnail_key": None,
            "regions": regions,
        }],
    }


def detect_excel_fields(excel_bytes: bytes, suffix: str = ".xlsx") -> tuple[list[str], list[dict]]:
    """
    Detect fillable fields in an Excel file.
    Two detection modes:
      1. Cells containing {{placeholder}} → field with that name
      2. Empty cells immediately to the right of a non-empty label cell → field
    Already-filled cells are never detected as fields.
    Returns (field_order, field_positions).
    """
    wb = openpyxl.load_workbook(BytesIO(_to_xlsx(excel_bytes, suffix)))
    ws = wb.active
    sheet_idx = wb.sheetnames.index(ws.title)

    field_order: list[str] = []
    field_positions: list[dict] = []
    seen: set[str] = set()

    def add_field(
        name: str,
        row: int,
        col: int,
        *,
        placeholder: bool = False,
        label: str | None = None,
    ) -> None:
        orig = name
        if name in seen:
            base, n = name, 2
            while f"{base}_{n}" in seen:
                n += 1
            name = f"{base}_{n}"
        seen.add(name)
        field_order.append(name)
        resolved_label = label or (orig.replace("_", " ").title() if placeholder else _label_from_position(ws, row, col))
        field_positions.append({
            "name": name,
            "label": resolved_label,
            "row": row,
            "col": col,
            "sheet": sheet_idx,
            "placeholder": placeholder,
        })

    for row in ws.iter_rows():
        for cell in row:
            # MergedCell proxies are not real cells — they belong to another cell's
            # merge range and their value is always None. Skip them entirely to avoid
            # detecting them as empty fields and then overwriting the merge master.
            if isinstance(cell, MergedCell):
                continue

            val = cell.value

            # Mode 1: cell contains {{placeholder}}
            if isinstance(val, str):
                match = PLACEHOLDER_RE.search(val)
                if match:
                    add_field(match.group(1), cell.row, cell.column, placeholder=True)
                    continue

            # Mode 2: empty cell to the right of a labelled cell
            is_empty = val is None or (isinstance(val, str) and not val.strip())
            if is_empty and cell.column > 1:
                left = ws.cell(row=cell.row, column=cell.column - 1)
                # Also skip if the left cell is a MergedCell proxy
                if isinstance(left, MergedCell):
                    continue
                left_val = left.value
                if left_val and isinstance(left_val, str) and _is_label(left_val):
                    label = left_val.strip().rstrip(":")
                    add_field(_make_field_name(label), cell.row, cell.column, label=label)

    return field_order, field_positions


def fill_excel(
    excel_bytes: bytes,
    values: dict[str, str],
    field_positions: list[dict],
    suffix: str = ".xlsx",
) -> bytes:
    """Fill detected fields in an Excel file and return .xlsx bytes."""
    wb = openpyxl.load_workbook(BytesIO(_to_xlsx(excel_bytes, suffix)))
    pos_by_name = {p["name"]: p for p in field_positions}

    for name, value in values.items():
        if not value or not value.strip():
            continue
        pos = pos_by_name.get(name)
        if not pos:
            continue
        if "row" not in pos or "col" not in pos or "sheet" not in pos:
            continue
        ws = wb.worksheets[pos["sheet"]]
        cell = _resolve_cell(ws, pos["row"], pos["col"])
        if pos.get("placeholder") and isinstance(cell.value, str):
            cell.value = cell.value.replace(f"{{{{{name}}}}}", value.strip())
        else:
            cell.value = value.strip()

    out = BytesIO()
    wb.save(out)
    return out.getvalue()


def excel_to_pdf(excel_bytes: bytes, suffix: str = ".xlsx") -> bytes:
    """Convert a filled Excel file to PDF using LibreOffice."""
    with tempfile.TemporaryDirectory() as tmpdir:
        tmp = Path(tmpdir)
        excel_path = tmp / f"form{suffix}"
        excel_path.write_bytes(excel_bytes)

        soffice = (
            shutil.which("soffice")
            or "/Applications/LibreOffice.app/Contents/MacOS/soffice"
        )
        subprocess.run(
            [soffice, "--headless", "--convert-to", "pdf", "--outdir", str(tmp), str(excel_path)],
            check=True,
            capture_output=True,
        )

        return (tmp / "form.pdf").read_bytes()
